import pandas as pd
import numpy as np
import pyautogui as py
import time
import openpyxl as op
import pptx
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import xlrd
import types
import os, glob
#packages

#
end = 1
while end == 1:
        monarch = input('Enter your monarch file without file tag:')#Takes user input about which file to use for monarch stats
        file = pd.read_excel(r'C:\Users\bomtved\Desktop\Press Reports\March 2023/'+monarch+'.xlsx')
        file = file.to_csv(r'C:\Users\bomtved\Desktop\Press Reports\March 2023/'+monarch+'.csv')
        file = pd.read_csv(r'C:\Users\bomtved\Desktop\Press Reports\March 2023/'+monarch+'.csv', index_col=[0])
        writer = pd.ExcelWriter(monarch+'.xlsx')
        file.to_excel(writer, index=False)
        writer.close()
        wb = load_workbook(monarch+'.xlsx')#Loads user input(must be xlsx file)
        fb = load_workbook('WCA Conversion.xlsx')#Base excel conversion sheet that creates visual for operator stats
        ws = wb.active#Loads base sheet of loaded file
        opps = ['Cletus Lee Landsteiner','Carlson, Todd M.','Strunk, Barry G',
        'Kollasch, Dennis L','Walter Church','Rock, Tim','McClay, Gene D.','Chavez, Abel',
        'Johnson, Charles D','Lonnie D. Pietsch','Marcus J. Schwartz','Wylie Morgan',
        'Bode, Chris','Michels, Troy A','Blom, Timothy L','Hanson, Rick','Thomas Koehn','Lundberg, Kenneth L'
        ,'Tiach Jiech War','Billy Crawford']#Possible operator names found in Monarch(must add any new operators)
        deps = ['Sanden 07 992-18" FixedSize',
                'Sanden 03 893-27" 1200',
                'Sanden 04  910-27" 1200',
                'Sanden 2',
                'Apollo 07',
                'Webcom 12',
                'Webcom 15',
                'Webcom 16',
                'Webcom 17',
                'Webcom 18',
                'Digital HS Inkjet',
                'Comco 02 Direct Response',
                'Sanden 06 Specialty 953-27"',
                'Apollo 05 - Navitor',
                'Webcom 14']#All the possible department numbers(must add any new departments)
        depar = []#Final department stats
        oppos = []#Final operator name for stats
        stats = []#Final stats
        clearall = 0#Rows are deleted until reaching determined number
        place = 0#Determines operator name and department place in lists
        allclear = 0#Switches between 0 and 1 based on whether or not rows need to be deleted
        currentdep = []#Determines which department operator is found in
#
#Creates new lists in which each of the operators will be placed into based on department worked in
        s7 = []
        s3and4 = []
        s2 = []
        a7 = []
        w12 = []
        w15 = []
        w16and17 = []
        w18 = []
        i1 = []
#
#Finding rows that do not pertain to main press operation (support, slab, etc)
        for i in range(1, ws.max_row+1):
      
                for j in range(1, 5):
                        cell = ws.cell(row=i, column=j)
                        if allclear == 1:
                                break

                        if cell.value =='292':
                                clearall = i
                                allclear = 1
                                break
                
                        elif cell.value =='299':
                                clearall = i
                                allclear = 1
                                break
                
                        elif cell.value =='29Q':
                                clearall = i
                                allclear = 1
                                break
#
#Delete rows that do not pertain to main press operations         
        while clearall != ws.max_row:
                        ws.delete_rows(ws.max_row)
#
#Lists every operator found and then places them in order into another list
        for i in range(1, ws.max_row+1):
      
                for j in range(1, 5):
                        cell = ws.cell(row=i, column=j)
                        
                        for op in opps:
                                if cell.value == op:
                                        if cell.value == 'Cletus Lee Landsteiner':
                                                oppos.insert(1000, 'Lee Landsteiner')
                                        else:
                                                oppos.insert(1000,cell.value)
#
#Determines department that employee is found under then adds the department into another list
        for i in range(1, ws.max_row+1):
      
                for j in range(1, 5):
                        cell = ws.cell(row=i, column=j)

                        for dep in deps:
                                if cell.value == dep:
                                        if len(currentdep) == 0:
                                                currentdep.append(cell.value)
                                        if currentdep[0] != cell.value:
                                                currentdep.clear()
                                                currentdep.append(cell.value)

                        if cell.value == 'Employee Total':
                                depar.insert(100,currentdep[0])
#
#Recodes all departments into their respective names rather than department #
        for i in range(0,len(depar)):
                if depar[i] == 'Comco 02 Direct Response':
                        depar[i] = 'Comco 2'
                if depar[i] == 'Sanden 07 992-18" FixedSize':
                        depar[i] = 'Sanden 7'
                if depar[i] == 'Sanden 03 893-27" 1200':
                        depar[i] = 'Sanden 3'
                if depar[i] == 'Sanden 04  910-27" 1200':
                        depar[i] = 'Sanden 4'
                if depar[i] == 'Sanden 2':
                        depar[i] = 'Sanden 2'
                if depar[i] == 'Apollo 07':
                        depar[i] = 'Apollo 7'
                if depar[i] == 'Webcom 12':
                        depar[i] = 'Webcom 12'
                if depar[i] == 'Webcom 15':
                        depar[i] = 'Webcom 15'
                if depar[i] == 'Webcom 16':
                        depar[i] = 'Webcom 16'
                if depar[i] == 'Webcom 17':
                        depar[i] = 'Webcom 17'
                if depar[i] == 'Webcom 18':
                        depar[i] = 'Webcom 18'
                if depar[i] == 'Digital HS Inkjet':
                        depar[i] = 'Digital HS Inkjet'
#
#Creates final list for each operator determining their department, name, hours worked, gross feet ran, net feet ran and waste feet ran
        for i in range(1, ws.max_row+1):
      
                for j in range(1, ws.max_column+1):
                        cell = ws.cell(row=i, column=j)
                        if cell.value == 'Employee Total':
                                hours = ws.cell(row=i, column = j+1)
                                gross = ws.cell(row=i, column = j+2)
                                net = ws.cell(row=i, column = j+3)
                                waste = ws.cell(row=i, column = j+4)
                                stats.insert(1000,[depar[place],oppos[place],hours.value,gross.value,net.value,waste.value])
                                place = place+1
#
#Seperates final list into each department
        for i in stats:
                if i[0] == 'Sanden 7':
                        s7.insert(1000, i)
                if i[0] == 'Sanden 3':
                        s3and4.insert(1000,i)
                if i[0] == 'Sanden 4':
                        s3and4.insert(1000,i)
                if i[0] == 'Sanden 2':
                        s2.insert(1000,i)
                if i[0] == 'Apollo 7':
                        a7.insert(1000,i)
                if i[0] == 'Webcom 12':
                        w12.insert(1000,i)
                if i[0] == 'Webcom 15':
                        w15.insert(1000,i)
                if i[0] == 'Webcom 16':
                        w16and17.insert(1000,i)
                if i[0] == 'Webcom 17':
                        w16and17.insert(1000,i)
                if i[0] == 'Webcom 18':
                        w18.insert(1000,i)
                if i[0] == 'Digital HS Inkjet':
                        i1.insert(1000,i)
#
#Inserts operator 'stats' into final excel sheet based on department
        opcol = 1
#sanden 7 sheet
        sanden7 = fb['Sanden 7']
        san7 = 2
        for op in s7:
                for i in op:
                        sanden7.cell(row=san7,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                san7 = san7+1
#sanden 3 and 4 sheet
        sanden34 = fb['Sanden 3&4']
        san34 = 4
        for op in s3and4:
                for i in op:
                        sanden34.cell(row=san34,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                san34 = san34+1
#sanden 2 sheet
        sanden2 = fb['Sanden 2']
        san2 = 2
        for op in s2:
                for i in op:
                        sanden2.cell(row=san2,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                san2 = san2+1
#Apollo 7 sheet
        apollo7 = fb['Apollo 7']
        apo7 = 2
        for op in a7:
                for i in op:
                        apollo7.cell(row=apo7,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                apo7 = apo7+1
#Webcom 12 sheet
        webcom12 = fb['Webcom 12']
        web12 = 2
        for op in w12:
                for i in op:
                        webcom12.cell(row=web12,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                web12 = web12+1
#Webcom 15 sheet
        webcom15 = fb['Webcom 15']
        web15 = 2
        for op in w15:
                for i in op:
                        webcom15.cell(row=web15,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                web15 = web15+1
#Webcom 16 and 17 sheet
        webcom1617 = fb['Webcom 16&17']
        web1617 = 4
        for op in w16and17:
                for i in op:
                        webcom1617.cell(row=web1617,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                web1617 = web1617+1
#Webcom 18 sheet
        webcom18 = fb['Webcom 18']
        web18 = 2
        for op in w18:
                for i in op:
                        webcom18.cell(row=web18,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                web18 = web18+1
#Digital HS Inkjet sheet
        inkjet = fb['Digital HS Inkjet']
        ink1 = 2
        for op in i1:
                for i in op:
                        inkjet.cell(row=ink1,column=opcol).value = i
                        opcol = opcol+1
                opcol = 1
                ink1 = ink1+1
#
#Empties the path's folder so wca file is easier to find
        dir = 'C:/Users/bomtved/Desktop/Press Reports/Daily'
        for file in os.scandir(dir):
            os.remove(file.path)
#Saves final file(print and sleep statements are added for dramatic effect)
        fb.save('C:/Users/bomtved/Desktop/Press Reports/Daily/'+monarch+'WCA'+'.xlsx')
        print('. . .')
        time.sleep(1)
        print('WCA Completed')
        time.sleep(1)
        end = 2
        while end == 2:
                retry = input('Would you like to load a different file?'+'\n'+'Type yes or no:')
                if retry == 'yes':
                        end = 1
                elif retry == 'no':
                        end = 3
                else:
                        print('Invalid response')
                        end = 2

        
                
